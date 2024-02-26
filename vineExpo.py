import functools
import time

import openpyxl
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup as bs
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook

f_path = '/Users/trishika/PycharmProjects/pythonProject/data/vine/wine.xlsx'
workbook = openpyxl.load_workbook(f_path)
sheet = workbook['data']
sheet1 = workbook['page']

options = Options()
options.add_argument('--headless')
options.add_argument('--disable-gpu')
options.add_argument("--user-agent= Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.75 Safari/537.36")

visited_urls = set()
for row in sheet1.iter_rows(min_row=1, max_col=1, max_row=sheet1.max_row, values_only=True):
    visited_urls.add(row[0])

last_visited_url = 0


def retry(max_retries, delay=1):
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            for _ in range(max_retries):
                try:
                    result = func(*args, **kwargs)
                    return result
                except Exception as e:
                    print(f"Retrying after error: {e}")
                    time.sleep(delay)
            raise Exception(f"Max retries reached for {func.__name__}")

        return wrapper

    return decorator


@retry(max_retries=3)
def scrape_data():
    global last_visited_url
    driver = webdriver.Chrome(options=options)
    driver.set_window_size(1200, 600)
    try:
        driver.get("https://wineparis-vinexpo.com/newfront/marketplace/exhibitors?pageNumber=1&limit=60")
        last_visited_url = driver.current_url
    except Exception:
        raise Exception('Error launching browser')
    scrape_pg(driver)


@retry(max_retries=3)
def scrape_pg(driver):
    global n, last_visited_url
    while True:

        try:
            WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#__next > div.MuiBox-root.css-g9qx4c > div.MuiBox-root.css-1roqr68 > main > div > div > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-md-8.MuiGrid-grid-lg-9.css-vhv0fi > div > div.MuiBox-root.css-0 > div.MuiGrid-root.MuiGrid-container.MuiGrid-spacing-xs-2.css-1fxdrvb ")))
            print(driver.current_url)
            # Set the scroll increment (adjust as needed)
            scroll_increment = 600

            # Initial scroll position
            scroll_position = 0

            # Get the total height of the page
            total_height = driver.execute_script(
                "return Math.max( document.body.scrollHeight, document.body.offsetHeight, document.documentElement.clientHeight, document.documentElement.scrollHeight, document.documentElement.offsetHeight);")

            while scroll_position < total_height:
                # Scroll down by the increment
                driver.execute_script(f"window.scrollBy(0, {scroll_increment});")

                total_height = driver.execute_script(
                    "return Math.max( document.body.scrollHeight, document.body.offsetHeight, document.documentElement.clientHeight, document.documentElement.scrollHeight, document.documentElement.offsetHeight);")

                # Wait for a short duration to allow content to load (adjust as needed)
                time.sleep(1)

                # Update the scroll position
                scroll_position += scroll_increment

            ele1 = driver.find_elements(By.CSS_SELECTOR, '#__next > div.MuiBox-root.css-g9qx4c > div.MuiBox-root.css-1roqr68 > main > div > div > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-md-8.MuiGrid-grid-lg-9.css-vhv0fi > div > div.MuiBox-root.css-0 > div.MuiGrid-root.MuiGrid-container.MuiGrid-spacing-xs-2.css-1fxdrvb > div')
            print(f'{len(ele1)}page start')
            WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#__next > div.MuiBox-root.css-g9qx4c > div.MuiBox-root.css-1roqr68 > main > div > div > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-md-8.MuiGrid-grid-lg-9.css-vhv0fi > div > div.MuiBox-root.css-0 > div.MuiGrid-root.MuiGrid-container.MuiGrid-spacing-xs-2.css-1fxdrvb > div:nth-child(1) > div > div.MuiBox-root.css-cduqi1 > div.MuiBox-root.css-ob69bz > a > div > div")))
            WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#__next > div.MuiBox-root.css-g9qx4c > div.MuiBox-root.css-1roqr68 > main > div > div > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-md-8.MuiGrid-grid-lg-9.css-vhv0fi > div > div.MuiBox-root.css-0 > div.MuiGrid-root.MuiGrid-container.MuiGrid-spacing-xs-2.css-1fxdrvb > div:nth-child("+str(len(ele1))+") > div > div.MuiBox-root.css-cduqi1 > div.MuiBox-root.css-ob69bz > a > div > div")))
            btn = '#__next > div.MuiBox-root.css-g9qx4c > div.MuiBox-root.css-1roqr68 > main > div > div > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-md-8.MuiGrid-grid-lg-9.css-vhv0fi > div > div.MuiBox-root.css-0 > div.MuiBox-root.css-1ek3hjv > nav > ul > li:nth-child(9) > button '
            WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, btn)))
            print('p1')
            button = driver.find_element(By.CSS_SELECTOR, btn)
            if driver.current_url in visited_urls:
                while driver.current_url in visited_urls:
                    print(f'page {driver.current_url} is already visited')
                    driver.execute_script("arguments[0].click();", button)
                    WebDriverWait(driver, 10).until(EC.url_changes(driver.current_url))

                print('going to next page')
                continue
            html = driver.page_source
        except Exception as e:
            print(e)
            driver.refresh()
            raise Exception("error loading page content")

        soup = bs(html, "html.parser")

        # titles = soup.select('#__next > div.MuiBox-root.css-g9qx4c > div.MuiBox-root.css-1roqr68 > main > div > div > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-md-8.MuiGrid-grid-lg-9.css-vhv0fi > div > div.MuiBox-root.css-0 > .MuiGrid-root.MuiGrid-container.MuiGrid-spacing-xs-2.css-1fxdrvb div')
        titles = soup.find_all('div', class_='MuiGrid-root MuiGrid-item MuiGrid-grid-xs-12 MuiGrid-grid-sm-6 MuiGrid-grid-md-4 MuiGrid-grid-lg-3 css-1m6ye84')
        print('page content loaded')
        print(len(titles))
        next_row = sheet.max_row + 1
        try:
            for indx, title in enumerate(titles):
                sele = 'div:nth-child(' + str(indx + 1) + ') > div > div.MuiBox-root.css-cduqi1 > div.MuiBox-root.css-ob69bz > a > div > div'
                name = title.select_one(sele).get_text()
                link = title.select_one('div:nth-child(' + str(indx + 1) + ') > div > div.MuiBox-root.css-cduqi1 > div.MuiBox-root.css-ob69bz > a')
                sheet.cell(row=next_row, column=1, value=name)
                sheet.cell(row=next_row, column=2, value='https://wineparis-vinexpo.com'+link['href'])
                print(f"{next_row}){name}--------{link['href']}")
                next_row += 1

        except Exception as e:
            print(e)
            driver.refresh()
            raise Exception("error scraping page content")
        visited_urls.add(driver.current_url)
        next_row1 = sheet1.max_row + 1
        sheet1.cell(row=next_row1, column=1, value=driver.current_url)
        print('Url added successfully')
        workbook.save(f_path)
        last_visited_url = driver.current_url
        if button.is_enabled():
            driver.execute_script("arguments[0].click();", button)
            WebDriverWait(driver, 10).until(EC.url_changes(driver.current_url))
        else:
            break

    # Close the Browser
    driver.close()


scrape_data()
